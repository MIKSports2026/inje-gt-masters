"""
scripts/generate_results_template.py
인제 GT 마스터즈 — 경기 결과 엑셀 입력 양식 생성 스크립트

실행: python3 scripts/generate_results_template.py
출력: public/templates/inje-gt-masters_results_template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'public', 'templates', 'inje-gt-masters_results_template.xlsx'
)

def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)


# ── Sheet 1: 경기결과 ───────────────────────────────────────────
def build_results_sheet(wb):
    ws = wb.create_sheet('경기결과')

    # 컬럼 너비
    col_widths = [8, 10, 16, 12, 12, 14, 8, 14, 12, 12, 8, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 헤더 행 (1행)
    headers = ['순위', '차량번호', '팀명', '드라이버1', '드라이버2',
               '차량', '랩수', '총시간', '갭', '패스티스트', '포인트', '상태']
    ws.row_dimensions[1].height = 28

    hdr_fill = make_fill('111111')
    hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = center_align()
        cell.border = thin_border()

    # 예시 행 데이터 (2~4행)
    sample_rows = [
        [1, '7',  '팀스피드웨이', '홍길동', '김철수', '아반떼 N',   25, '0:45:12.345', '',        '1:45.678', 25, 'classified'],
        [2, '13', '레이싱랩',     '이영희', '',       '벨로스터 N', 25, '0:45:18.901', '+6.556',  '1:46.123', 18, 'classified'],
        [3, '22', '오토클럽',     '박민수', '정수연', '아반떼 N',   23, '',             '+2 laps', '1:47.890',  0, 'dnf'],
    ]

    sample_fill = make_fill('F0F0F0')
    sample_font = Font(name='Arial', italic=True, size=10, color='888888')

    for r_idx, row_data in enumerate(sample_rows, 2):
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val != '' else None)
            cell.fill  = sample_fill
            cell.font  = sample_font
            cell.alignment = center_align()
            cell.border = thin_border()

    # 주의 메시지 행 (5행, A5:L5 병합)
    ws.row_dimensions[5].height = 24
    ws.merge_cells('A5:L5')
    warn_cell = ws['A5']
    warn_cell.value = '⚠️ 위 3행은 예시입니다. 실제 입력 전 반드시 삭제 후 데이터를 입력하세요.'
    warn_cell.fill  = make_fill('FFF4E6')
    warn_cell.font  = Font(name='Arial', bold=True, size=10, color='B54708')
    warn_cell.alignment = left_align()
    # 병합 셀 외곽 테두리
    side = Side(style='thin', color='F59E0B')
    warn_cell.border = Border(left=side, right=side, top=side, bottom=side)


# ── Sheet 2: 작성가이드 ─────────────────────────────────────────
def build_guide_sheet(wb):
    ws = wb.create_sheet('작성가이드')

    # 컬럼 너비
    for col, w in zip('ABCDE', [12, 8, 22, 50, 22]):
        ws.column_dimensions[col].width = w

    # 헤더 행
    guide_headers = ['컬럼', '필수', '형식', '설명', '예시']
    ws.row_dimensions[1].height = 28
    hdr_fill = make_fill('111111')
    hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')

    for c, h in enumerate(guide_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = center_align()
        cell.border = thin_border()

    # 가이드 데이터
    guide_rows = [
        ('순위',       '필수', '정수',                    '1 이상의 양의 정수. 중복 불가',                                                                '1, 2, 3'),
        ('차량번호',   '선택', '텍스트',                  '차량에 부착된 번호 (숫자/문자 가능)',                                                            '7, 13, 22A'),
        ('팀명',       '필수', '텍스트',                  '참가 팀 공식 명칭',                                                                             '팀스피드웨이'),
        ('드라이버1',  '필수', '텍스트',                  '대표 드라이버 이름',                                                                            '홍길동'),
        ('드라이버2',  '선택', '텍스트',                  '공동 드라이버 이름. 1인 참가 시 공란',                                                          '김철수'),
        ('차량',       '선택', '텍스트',                  '차량 모델명',                                                                                   '아반떼 N'),
        ('랩수',       '선택', '정수',                    '완주한 랩 수',                                                                                  '25'),
        ('총시간',     '선택', '시:분:초.밀리',           '완주 총 시간. DNF/DNS/DSQ는 공란',                                                              '0:45:12.345'),
        ('갭',         '선택', '텍스트',                  "선두와의 차이. 1위는 공란, 랩차는 '+N laps'",                                                   '+6.556, +2 laps'),
        ('패스티스트', '선택', '분:초.밀리',              '최고 랩 타임',                                                                                  '1:45.678'),
        ('포인트',     '선택', '정수',                    '획득 포인트. 0도 명시 권장',                                                                    '25, 18, 15'),
        ('상태',       '선택', '영문 소문자 4가지 중 하나',
         'classified=완주 / dnf=Did Not Finish / dns=Did Not Start / dsq=Disqualified. 생략 시 classified',
         'classified'),
    ]

    req_font    = Font(name='Arial', bold=True, size=10, color='C81E1E')
    opt_font    = Font(name='Arial', size=10, color='444444')
    normal_font = Font(name='Arial', size=10, color='222222')
    alt_fill    = make_fill('F8F8F8')

    for r_idx, row_data in enumerate(guide_rows, 2):
        ws.row_dimensions[r_idx].height = 40
        fill = alt_fill if r_idx % 2 == 0 else None

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border()

            # 필수/선택 컬럼 (B열) 색상
            if c_idx == 2:
                cell.font = req_font if val == '필수' else opt_font
                cell.alignment = center_align(wrap=True)
            else:
                cell.font = normal_font


def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = Workbook()
    # 기본 시트 제거
    del wb[wb.sheetnames[0]]

    build_results_sheet(wb)
    build_guide_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f'생성 완료: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
